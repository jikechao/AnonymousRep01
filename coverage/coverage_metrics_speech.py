#encoding=utf-8
from keras.datasets import cifar10, cifar100, fashion_mnist
import keras
import openpyxl
import os
import argparse
import numpy as np

from keras.datasets import mnist
from keras.models import load_model, Model
from sa import fetch_dsa, fetch_lsa, get_sc
from utils import *

import datetime
from keras.applications import vgg19,resnet50
import pickle

import deepspeech_utils as ds_utils
from deepspeech_utils import DSDataUtils
from deepspeech_utils import dataGen_mfcc_ctc
from deepspeech_text import Alphabet
import keras.backend as K

def get_score(x_test, y_test, model, column):
    total = 0
    diff = 0
    correct = 0
    alphabet = Alphabet(os.path.join(basedir,'data/alphabet.txt'))

    y_true = [[alphabet._label_to_str[y] for y in x] for x in y_test]
    y_true = [''.join(x).strip() for x in y_true]

    y_pred = model.predict(x_test, verbose=2)
    input_len = np.ones(y_pred.shape[0]) * y_pred.shape[1]

    y_pred = K.get_value(K.ctc_decode(y_pred, input_length=input_len)[0][0])
    y_pred = [[alphabet._label_to_str[y] for y in x if y >= 0] for x in y_pred]
    y_pred = [''.join(x).strip() for x in y_pred]
    y_pred = [DSDataUtils.get_bestmatch_keywords_using_wer(x) for x in y_pred]

    for a, b in list(zip(y_true, y_pred)):
        total += 1
        if a == b:
            correct += 1

    print(correct / total)
    print(correct)
    print(total)

    sheet.cell(row=10, column=column).value = (1 - correct / total)
    return correct / total

if __name__=="__main__":
    start = datetime.datetime.now()

    basedir = os.path.abspath(os.path.dirname(__file__))

    parser = argparse.ArgumentParser()
    parser.add_argument("--count", "-count", help="the index of group", type=int, default=0)
    parser.add_argument("--size", "-size", help="the number of each group", type=int, default=1000)
    parser.add_argument("--d", "-d", help="Dataset", type=str, default="speech")
    parser.add_argument(
        "--lsa", "-lsa", help="Likelihood-based Surprise Adequacy", action="store_true"
    )
    parser.add_argument(
        "--dsa", "-dsa", help="Distance-based Surprise Adequacy", action="store_true"
    )
    parser.add_argument(
        "--target",
        "-target",
        help="Target input set (test or adversarial set)",
        type=str,
        default="cw",
    )
    parser.add_argument(
        "--save_path", "-save_path", help="Save path", type=str, default="./tmp/"
    )
    parser.add_argument(
        "--batch_size", "-batch_size", help="Batch size", type=int, default=128
    )
    parser.add_argument(
        "--var_threshold",
        "-var_threshold",
        help="Variance threshold",
        type=int,
        default=1e-5,
    )
    parser.add_argument(
        "--upper_bound", "-upper_bound", help="Upper bound", type=int
    )
    parser.add_argument(
        "--n_bucket",
        "-n_bucket",
        help="The number of buckets for coverage",
        type=int,
        default=1000,
    )
    parser.add_argument(
        "--num_classes",
        "-num_classes",
        help="The number of classes",
        type=int,
        default=29,
    )
    parser.add_argument(
        "--is_classification",
        "-is_classification",
        help="Is classification task",
        type=bool,
        default=True,
    )
    parser.add_argument(
        "--lsa_lower",
        "-lsa_lower",
        help="Lower bound of lsa",
        type=int,
        default=0,
    )
    parser.add_argument(
        "--lsa_upper",
        "-lsa_upper",
        help="Upper bound of lsa",
        type=int,
        default=5000,
    )
    parser.add_argument(
        "--dsa_lower",
        "-dsa_lower",
        help="Lower bound of dsa",
        type=int,
        default=0,
    )
    parser.add_argument(
        "--dsa_upper",
        "-dsa_upper",
        help="Upper bound of dsa",
        type=int,
        default=4,
    )
    args = parser.parse_args()
    print(args)

    if os.path.exists(os.path.join(basedir, "result", "{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target))):
        workbook = openpyxl.load_workbook(os.path.join(basedir, "result", "{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target)))
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active





    batch_size = 51776
    trGen = dataGen_mfcc_ctc(ds_utils.trfile, batchSize=batch_size)
    [x_train, y_train, _, _], _ = next(DSDataUtils.gen_ctc_byclass(trGen))


    batch_size = 6471
    testGen = dataGen_mfcc_ctc(ds_utils.testfile, batchSize=batch_size)
    [x_test1, y_test1, _, _], _ = next(DSDataUtils.gen_ctc_byclass(testGen))

    advfile = os.path.join(basedir,'data/speechcmd_adv.csv')
    batch_size = 6824
    advGen = dataGen_mfcc_ctc(advfile, batchSize=batch_size)
    [x_test2, y_test2, _, _], _ = next(DSDataUtils.gen_ctc_byclass(advGen))

    model = load_model("./model/deepspeech-speech-command_origin.h5")
    model.summary()

    layer_names = ["dense_1"]

    import pickle
    with open('labels-500groups-{}samples.pk'.format(args.size), 'rb') as fp:
        labels = pickle.load(fp)

    X_test2 = np.vstack((x_test1[labels['normal'][args.count]], x_test2[labels['adv'][args.count]]))
    Y_test2 = np.vstack((y_test1[labels['normal'][args.count]], y_test2[labels['adv'][args.count]]))

    print(args.count)
    sheet.cell(row=1, column=args.count+1).value = args.count
    sheet.cell(row=9, column=args.count+1).value = labels['p'][args.count]

    # sheet.cell(row=9, column=args.count + 1).value = 0
    # np.random.seed(args.count+1)
    # labels = np.random.randint(low=0, high=5000,size = 1000)
    # X_test2 = x_test1[labels]
    # Y_test2 = y_test1[labels]

    # X_test2 = np.vstack((x_test1, x_test2))
    # Y_test2 = np.vstack((y_test1, y_test2))

    print(X_test2.shape)
    print(Y_test2.shape)

    get_score(X_test2,Y_test2,model,args.count+1)


    # print(x_train.shape)
    # X_test2 = []
    target_lsa = fetch_lsa(model=model, x_train=x_train, x_target=X_test2, target_name=args.target, layer_names=layer_names, args=args)
    target_cov = get_sc(lower=args.lsa_lower, upper=args.lsa_upper, k=args.n_bucket, sa=target_lsa)

    target_lsa = [x for x in target_lsa if str(x) != 'nan']
    # sheet.cell(row=2, column=args.count+1).value = np.max(target_lsa)
    # sheet.cell(row=3, column=args.count + 1).value = np.min(target_lsa)
    sheet.cell(row=4, column=args.count + 1).value = target_cov
    print(len(target_lsa))
    print("max", np.max(target_lsa))
    print("min", np.min(target_lsa))
    print(infog("{} coverage: ".format(args.target) + str(target_cov)))


    target_dsa = fetch_dsa(model, x_train, X_test2, args.target, layer_names, args)
    target_cov = get_sc(lower=args.dsa_lower, upper=args.dsa_upper, k=args.n_bucket, sa=target_dsa)

    # sheet.cell(row=5, column=args.count+1).value = np.max(target_dsa)
    # sheet.cell(row=6, column=args.count + 1).value = np.min(target_dsa)
    sheet.cell(row=7, column=args.count + 1).value = target_cov
    print(infog("{} coverage: ".format(args.target) + str(target_cov)))
    print(len(target_dsa))
    print("max", np.max(target_dsa))
    print("min", np.min(target_dsa))

    workbook.save(os.path.join(basedir, "result","{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target)))
    workbook.close()


    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)