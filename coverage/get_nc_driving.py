from tools.utils4 import NeuronCoverage,K_SectionNeuronCoverage
from keras.datasets import cifar10, cifar100, fashion_mnist
import keras
import openpyxl
import os
import argparse

from keras.datasets import mnist
from keras.models import load_model, Model
from utils import *

import datetime

workbook = openpyxl.Workbook()
sheet = workbook.active

CLIP_MIN = -0.5
CLIP_MAX = 0.5

from driving_models import *

# input image dimensions
img_rows, img_cols = 100, 100
input_shape = (img_rows, img_cols, 3)

# define input tensor as a placeholder
input_tensor = Input(shape=input_shape)

# load multiple models sharing same input tensor
# model = Dave_orig(input_tensor=input_tensor, load_weights=True)
# model = Dave_norminit(input_tensor=input_tensor, load_weights=True)
model = Dave_dropout(input_tensor=input_tensor, load_weights=True)



def get_score(x_test, y_test, model, column, row):
    pred = model.predict(x_test).reshape(-1)
    y_test = np.squeeze(y_test)
    true_acc = np.sum(np.square(pred - y_test)) / x_test.shape[0]
    print('Test accuracy:', true_acc)

    sheet.cell(row=row, column=column).value = true_acc
    return true_acc

def load_data():
    path = os.path.join(basedir,'driving/testing/final_example.csv')
    temp = np.loadtxt(path, delimiter=',', dtype=np.str, skiprows=(1))
    names = list(temp[:, 0])
    test = []
    label = []
    for i in range(len(names)):
        n = names[i]
        path = 'driving/testing/center/' + n + '.jpg'
        path = os.path.join(basedir, path)
        test.append(preprocess_image(path))
        label.append(float(temp[i, 1]))
    test = np.array(test)
    test = test.reshape(test.shape[0], 100, 100, 3)
    label = np.array(label)
    return test, label

def add_light(temp, gradients):
    # import skimage
    temp = temp.reshape(temp.shape[0], -1)
    gradients = gradients.reshape(gradients.shape[0], -1)
    new_grads = np.ones_like(gradients)
    grad_mean = 500 * np.mean(gradients, axis=1)
    grad_mean = np.tile(grad_mean, temp.shape[1])
    grad_mean = grad_mean.reshape(temp.shape)
    temp = temp + 80 * new_grads * grad_mean
    temp = temp.reshape(temp.shape[0], 100, 100, 3)
    return temp

def add_black(temp, gradients):
    rect_shape = (30, 30)
    for i in range(temp.shape[0]):
        orig = temp[i].reshape(1, 100, 100, 3)
        grad = gradients[i].reshape(1, 100, 100, 3)
        start_point = (
            random.randint(0, grad.shape[1] - rect_shape[0]), random.randint(0, grad.shape[2] - rect_shape[1]))
        new_grads = np.zeros_like(grad)
        patch = grad[:, start_point[0]:start_point[
            0] + rect_shape[0], start_point[1]:start_point[1] + rect_shape[1]]
        new_grads[:, start_point[0]:start_point[0] + rect_shape[0],
                      start_point[1]:start_point[1] + rect_shape[1]] = -np.ones_like(patch)
        orig = orig + 100 * new_grads
        temp[i] = orig.reshape(100, 100, 3)
    return temp

if __name__=="__main__":
    start = datetime.datetime.now()

    basedir = os.path.abspath(os.path.dirname(__file__))

    parser = argparse.ArgumentParser()
    parser.add_argument("--size", "-size", help="the number of each group", type=int, default=1000)

    parser.add_argument(
        "--target",
        "-target",
        help="Target input set (test or adversarial set)",
        type=str,
        default="black",
    )

    args = parser.parse_args()
    print(args)

    test, label = load_data()
    print("data loaded!")

    x_test1 = test.copy()
    y_test1 = label.copy()

    if args.target == "black":
        temp = test.copy()
        pert = 1 * np.random.normal(size=x_test1.shape)
        for i in range(7):
            temp = add_black(temp, pert)

        x_test2 = temp.copy()
        y_test2 = label.copy()

    elif args.target == "light":
        temp = test.copy()
        pert = 1 * np.random.normal(size=x_test1.shape)
        temp = add_light(temp, pert)
        x_test2 = temp.copy()
        y_test2 = label.copy()

    y_test1 = np.expand_dims(y_test1, axis=1)
    y_test2 = np.expand_dims(y_test2, axis=1)

    import pickle
    with open('labels-500groups-{}samples.pk'.format(args.size), 'rb') as fp:
        labels = pickle.load(fp)

    with open("driving_dave_drop", "rb")as file:
        boundry = pickle.load(file)

    for count in range(500):
        p = labels['p'][count]
        p_100 = int(100 * p)
        p_500 = int(500 * p)

        start2 = datetime.datetime.now()

        X_test_100 = np.vstack((x_test1[labels['normal'][count][:100 - p_100]], x_test2[labels['adv'][count][:p_100]]))
        Y_test_100 = np.vstack((y_test1[labels['normal'][count][:100 - p_100]], y_test2[labels['adv'][count][:p_100]]))

        X_test_400 = np.vstack(
            (x_test1[labels['normal'][count][100 - p_100: 500 - p_500]], x_test2[labels['adv'][count][p_100: p_500]]))
        Y_test_400 = np.vstack(
            (y_test1[labels['normal'][count][100 - p_100: 500 - p_500]], y_test2[labels['adv'][count][p_100: p_500]]))

        X_test_500 = np.vstack((x_test1[labels['normal'][count][500 - p_500:]],
                                x_test2[labels['adv'][count][p_500:]]))
        Y_test_500 = np.vstack((y_test1[labels['normal'][count][500 - p_500:]],
                                y_test2[labels['adv'][count][p_500:]]))

        X_test2 = np.vstack((X_test_100, X_test_400))
        Y_test2 = np.vstack((Y_test_100, Y_test_400))

        s = get_score(X_test_100, Y_test_100, model, count + 1, 22)
        s = get_score(X_test2, Y_test2, model, count + 1, 23)

        X_test2 = np.vstack((X_test2, X_test_500))
        Y_test2 = np.vstack((Y_test2, Y_test_500))

        # X_test2 = np.vstack((x_test1[labels['normal'][count]], x_test2[labels['adv'][count]]))
        # Y_test2 = np.vstack((y_test1[labels['normal'][count]], y_test2[labels['adv'][count]]))

        print(count)
        # print(X_test2.shape)
        # print(Y_test2.shape)

        sheet.cell(row=1, column=count + 1).value = count
        sheet.cell(row=20, column=count + 1).value = labels['p'][count]

        # 单纯的adv或者normal，不混合
        # sheet.cell(row=20, column=count + 1).value = 0
        # labels = np.random.randint(low=0, high=5613,size = 1000)
        # X_test2 = x_test1[labels]
        # Y_test2 = y_test1[labels]

        get_score(X_test2,Y_test2,model,count+1, 24)

        coverage1 = NeuronCoverage(model=model)
        final_coverage_rate = coverage1.final_coverage(inputs=X_test2, threshold=0.5, K=1, b=boundry)

        sheet.cell(row=3, column=count + 1).value = final_coverage_rate[0]
        sheet.cell(row=4, column=count + 1).value = final_coverage_rate[1]
        sheet.cell(row=5, column=count + 1).value = final_coverage_rate[2]
        sheet.cell(row=6, column=count + 1).value = final_coverage_rate[3]

        sheet.cell(row=7, column=count + 1).value = final_coverage_rate[4]
        sheet.cell(row=8, column=count + 1).value = final_coverage_rate[5]
        sheet.cell(row=9, column=count + 1).value = final_coverage_rate[6]
        sheet.cell(row=10, column=count + 1).value = final_coverage_rate[7]

        sheet.cell(row=11, column=count + 1).value = final_coverage_rate[8]
        sheet.cell(row=12, column=count + 1).value = final_coverage_rate[9]
        sheet.cell(row=13, column=count + 1).value = final_coverage_rate[10]
        sheet.cell(row=14, column=count + 1).value = final_coverage_rate[11]
        print("------------")

        start3 = datetime.datetime.now()
        coverage2 = K_SectionNeuronCoverage(model=model)
        kmnc = coverage2.final_coverage(inputs=X_test2, k_section=1000, b=boundry)
        elapsed3 = (datetime.datetime.now() - start3)
        print("Time used: ", elapsed3)
        print("------------")
        sheet.cell(row=16, column=count + 1).value = kmnc[0]
        sheet.cell(row=17, column=count + 1).value = kmnc[1]
        sheet.cell(row=18, column=count + 1).value = kmnc[2]


        workbook.save(os.path.join(basedir, "driving_result","drop-{}samples-500groups-kmnc-{}.xlsx".format(args.size, args.target)))
        workbook.close()

        elapsed2 = (datetime.datetime.now() - start2)
        print("Time used: ", elapsed2)


    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)