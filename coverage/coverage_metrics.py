#encoding=utf-8
from keras.datasets import cifar10, cifar100, fashion_mnist
import keras
import openpyxl
import os
import argparse
import numpy as np

from keras.datasets import mnist
from keras.models import load_model, Model
from sa import fetch_dsa, fetch_lsa, get_sc
from utils import *

import datetime
from keras.applications import vgg19,resnet50
import pickle
# from driving_models import *

def get_score(x_test, y_test, model, column):
    Y_test = keras.utils.to_categorical(y_test, args.num_classes)

    score = model.evaluate(x_test, Y_test, verbose=0)

    print('Test accuracy:', score[1])
    print('预测错的数目：', len(x_test)*(1-score[1]))

    sheet.cell(row=10, column=column).value = (1 - score[1])
    return score

def get_adv_mnist(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_mnist_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_mnist_label.npy'.format(args.target))
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test,y

def get_adv_cifar10(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_cifar10_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_cifar10_label.npy'.format(args.target))
    x_test = np.load(image_path)
    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_svhn(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_svhn_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_svhn_label.npy'.format(args.target))
    x_test = np.load(image_path)
    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_cifar100(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_cifar100_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_cifar100_label.npy'.format(args.target))
    x_test = np.load(image_path)
    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_fashion(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_fashion_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_fashion_label.npy'.format(args.target))
    x_test = np.load(image_path)
    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_imagenet(**kwargs):
    for i in range(0, 1000, 100):
        image_path = os.path.join(basedir, 'data', 'imagenet_adv/{}_imagenet_image_{}.npy'.format(args.target, i))
        label_path = os.path.join(basedir, 'data', 'imagenet_adv/{}_imagenet_label_{}.npy'.format(args.target, i))
        x_test = np.load(image_path)
        y_test = np.load(label_path)

        y = []
        for temp in y_test:
            temp_list = []
            temp_list.append(int(temp))
            y.append(temp_list)
        y = np.array(y)

        if i == 0:
            X_test = x_test
            Y_test = y
        else:
            X_test = np.vstack((X_test, x_test))
            Y_test = np.vstack((Y_test, y))

    return X_test,Y_test

def get_score_driving(x_test, y_test, model, column):
    pred = model.predict(x_test).reshape(-1)
    y_test = np.squeeze(y_test)
    true_acc = np.sum(np.square(pred - y_test)) / x_test.shape[0]
    print('Test accuracy:', true_acc)

    sheet.cell(row=10, column=column).value = true_acc
    return true_acc

def load_data():
    path = os.path.join(basedir,'driving/testing/final_example.csv')
    temp = np.loadtxt(path, delimiter=',', dtype=np.str, skiprows=(1))
    names = list(temp[:, 0])
    test = []
    label = []
    for i in range(len(names)):
        n = names[i]
        path = 'driving/testing/center/' + n + '.jpg'
        path = os.path.join(basedir, path)
        test.append(preprocess_image(path))
        label.append(float(temp[i, 1]))
    test = np.array(test)
    test = test.reshape(test.shape[0], 100, 100, 3)
    label = np.array(label)
    return test, label

def load_driving_train():
    with open('image_root_d', 'rb') as file:
        image_list = pickle.load(file)
    with open('image_root_d_left', 'rb') as file:
        image_list2 = pickle.load(file)
    with open('image_root_d_right', 'rb') as file:
        image_list3 = pickle.load(file)
    image_list.extend(image_list2)
    image_list.extend(image_list3)
    print(len(image_list))
    train = []
    # label = []
    for path in image_list:
        train.append(preprocess_image(path))
        # label.append(float(temp[i, 1]))
    train = np.array(train)
    train = train.reshape(train.shape[0], 100, 100, 3)
    # label = np.array(label)
    return train

def load_imagenet_train():
    with open('image_root', 'rb') as file:
        image_list = pickle.load(file)
    image_list = image_list[950001:1000000]
    print(len(image_list))
    train = []
    # label = []
    for path in image_list:
        train.append(preprocess_image(path))
        # label.append(float(temp[i, 1]))
    train = np.array(train)
    train = train.reshape(train.shape[0], 224, 224, 3)
    # label = np.array(label)
    return train


def add_light(temp, gradients):
    temp = temp.reshape(temp.shape[0], -1)
    gradients = gradients.reshape(gradients.shape[0], -1)
    new_grads = np.ones_like(gradients)
    grad_mean = 500 * np.mean(gradients, axis=1)
    grad_mean = np.tile(grad_mean, temp.shape[1])
    grad_mean = grad_mean.reshape(temp.shape)
    temp = temp + 80 * new_grads * grad_mean
    temp = temp.reshape(temp.shape[0], 100, 100, 3)
    return temp

def add_black(temp, gradients):
    rect_shape = (30, 30)
    for i in range(temp.shape[0]):
        orig = temp[i].reshape(1, 100, 100, 3)
        grad = gradients[i].reshape(1, 100, 100, 3)
        start_point = (
            random.randint(0, grad.shape[1] - rect_shape[0]), random.randint(0, grad.shape[2] - rect_shape[1]))
        new_grads = np.zeros_like(grad)
        patch = grad[:, start_point[0]:start_point[
            0] + rect_shape[0], start_point[1]:start_point[1] + rect_shape[1]]
        new_grads[:, start_point[0]:start_point[0] + rect_shape[0],
                      start_point[1]:start_point[1] + rect_shape[1]] = -np.ones_like(patch)
        orig = orig + 100 * new_grads
        temp[i] = orig.reshape(100, 100, 3)
    return temp

if __name__=="__main__":
    start = datetime.datetime.now()

    basedir = os.path.abspath(os.path.dirname(__file__))

    parser = argparse.ArgumentParser()
    parser.add_argument("--count", "-count", help="the index of group", type=int, default=0)
    parser.add_argument("--size", "-size", help="the number of each group", type=int, default=1000)
    parser.add_argument("--d", "-d", help="Dataset", type=str, default="mnist")
    parser.add_argument(
        "--lsa", "-lsa", help="Likelihood-based Surprise Adequacy", action="store_true"
    )
    parser.add_argument(
        "--dsa", "-dsa", help="Distance-based Surprise Adequacy", action="store_true"
    )
    parser.add_argument(
        "--target",
        "-target",
        help="Target input set (test or adversarial set)",
        type=str,
        default="cw",
    )
    parser.add_argument(
        "--save_path", "-save_path", help="Save path", type=str, default="./tmp/"
    )
    parser.add_argument(
        "--batch_size", "-batch_size", help="Batch size", type=int, default=128
    )
    parser.add_argument(
        "--var_threshold",
        "-var_threshold",
        help="Variance threshold",
        type=int,
        default=1e-5,
    )
    parser.add_argument(
        "--upper_bound", "-upper_bound", help="Upper bound", type=int
    )
    parser.add_argument(
        "--n_bucket",
        "-n_bucket",
        help="The number of buckets for coverage",
        type=int,
        default=1000,
    )
    parser.add_argument(
        "--num_classes",
        "-num_classes",
        help="The number of classes",
        type=int,
        default=10,
    )
    parser.add_argument(
        "--is_classification",
        "-is_classification",
        help="Is classification task",
        type=bool,
        default=True,
    )
    parser.add_argument(
        "--lsa_lower",
        "-lsa_lower",
        help="Lower bound of lsa",
        type=int,
        default=0,
    )
    parser.add_argument(
        "--lsa_upper",
        "-lsa_upper",
        help="Upper bound of lsa",
        type=int,
        default=5000,
    )
    parser.add_argument(
        "--dsa_lower",
        "-dsa_lower",
        help="Lower bound of dsa",
        type=int,
        default=0,
    )
    parser.add_argument(
        "--dsa_upper",
        "-dsa_upper",
        help="Upper bound of dsa",
        type=int,
        default=4,
    )
    args = parser.parse_args()
    print(args)

    if os.path.exists(os.path.join(basedir, "result", "{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target))):
        workbook = openpyxl.load_workbook(os.path.join(basedir, "result", "{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target)))
    else:
        workbook = openpyxl.Workbook()

    sheet = workbook.active



    if args.d == "mnist":
        (x_train, y_train), (x_test1, y_test) = mnist.load_data()

        (x_test2, y_test2) = get_adv_mnist()
        x_train = x_train.reshape(-1, 28, 28, 1)
        x_test1 = x_test1.reshape(-1, 28, 28, 1)

        y_test1 = np.expand_dims(y_test,axis=1)

        # Load pre-trained model.
        model = load_model("./model/model_mnist.hdf5")
        model.summary()
        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

        layer_names = ["activation_14"]
        # layer_names = ["conv2d_2"]

    elif args.d == "cifar":
        (x_train, y_train), (x_test1, y_test1) = cifar10.load_data()

        (x_test2, y_test2) = get_adv_cifar10()

        # model = load_model("./model/cifar10-vgg16_model_alllayers.h5")
        model = load_model("./model/model_cifar10.h5")
        model.summary()

        x_train = x_train.reshape(x_train.shape[0], 32, 32, 3)
        x_test1 = x_test1.reshape(x_test1.shape[0], 32, 32, 3)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

        # layer_names = ["block3_pool"]
        # layer_names = ["dense_1"]
        layer_names = ["activation_19"]

    elif args.d == "svhn":
        (x_train, y_train), (x_test1, y_test1) = SVNH_DatasetUtil.load_data()
        (x_test2, y_test2) = get_adv_svhn()

        model = load_model("./model/model_svhn.hdf5")
        model.summary()
        layer_names = ["activation_2"]

        y_test1 = np.argmax(y_test1, axis=1)
        y_test1 = y_test1.reshape(-1, 1)

    elif args.d == "fashion":
        (x_train, y_train), (x_test1, y_test1) = fashion_mnist.load_data()

        (x_test2, y_test2) = get_adv_fashion()

        x_train = x_train.reshape(-1, 28, 28, 1)
        x_test1 = x_test1.reshape(-1, 28, 28, 1)
        y_train = y_train.reshape(-1, 1)
        y_test1 = y_test1.reshape(-1, 1)

        model = load_model("./model/model_fashion.hdf5")
        model.summary()

        layer_names = ["activation_2"]

    elif args.d == "cifar100":
        (x_train, y_train), (x_test1, y_test1) = cifar100.load_data()

        (x_test2, y_test2) = get_adv_cifar100()

        model = load_model("./model/model_cifar100.h5")
        model.summary()

        layer_names = ["activation_28"]

        x_train = x_train.reshape(x_train.shape[0], 32, 32, 3)
        x_test1 = x_test1.reshape(x_test1.shape[0], 32, 32, 3)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255
    elif args.d == "imagenet":

        # x_train = load_imagenet_train()
        # print(x_train.shape)

        x_train = []

        data_path = os.path.join(basedir, 'data', "imagenet-val-5000.npz")
        # data_path = os.path.join(basedir, 'data', "imagenet.npz")
        data = np.load(data_path)
        x_test1, y_test1 = data['x_test'], data['y_test']

        x_test1 = vgg19.preprocess_input(x_test1)
        # x_test1 = resnet50.preprocess_input(x_test1)

        (x_test2, y_test2) = get_adv_imagenet()
        x_test2 = vgg19.preprocess_input(x_test2)
        # x_test2 = resnet50.preprocess_input(x_test2)
        model = vgg19.VGG19(weights='imagenet')
        # model = resnet50.ResNet50(weights='imagenet')
        model.summary()

        model.compile(loss="categorical_crossentropy", optimizer="adadelta", metrics=["accuracy"])
        y_test1 = y_test1.reshape(y_test1.shape[0],1)
        layer_names = ["block5_conv4"]
        # layer_names = ["activation_48"]
    elif args.d == "driving_ori" or args.d == "driving_drop":
        from driving_models import *

        # input image dimensions
        img_rows, img_cols = 100, 100
        input_shape = (img_rows, img_cols, 3)

        # define input tensor as a placeholder
        input_tensor = Input(shape=input_shape)

        # load multiple models sharing same input tensor
        if args.d == "driving_ori":
            model = Dave_orig(input_tensor=input_tensor, load_weights=True)
        elif args.d == "driving_drop":
            model = Dave_dropout(input_tensor=input_tensor, load_weights=True)

        model.summary()

        layer_names = ["block1_conv2"]

        test, label = load_data()
        print("data loaded!")

        # x_train = load_driving_train()
        x_train = []

        x_test1 = test.copy()
        y_test1 = label.copy()

        if args.target == "black":
            temp = test.copy()
            pert = 1 * np.random.normal(size=x_test1.shape)
            for i in range(7):
                temp = add_black(temp, pert)

            x_test2 = temp.copy()
            # test = temp.copy()
            y_test2 = label.copy()

        elif args.target == "light":
            temp = test.copy()
            pert = 1 * np.random.normal(size=x_test1.shape)
            temp = add_light(temp, pert)
            x_test2 = temp.copy()
            # test = temp.copy()
            y_test2 = label.copy()

        y_test1 = np.expand_dims(y_test1, axis=1)
        y_test2 = np.expand_dims(y_test2, axis=1)


    import pickle
    with open('labels-1000groups-{}samples.pk'.format(args.size), 'rb') as fp:
        labels = pickle.load(fp)

    X_test2 = np.vstack((x_test1[labels['normal'][args.count]], x_test2[labels['adv'][args.count]]))
    Y_test2 = np.vstack((y_test1[labels['normal'][args.count]], y_test2[labels['adv'][args.count]]))

    print(args.count)
    sheet.cell(row=1, column=args.count+1).value = args.count
    sheet.cell(row=9, column=args.count+1).value = labels['p'][args.count]

    # sheet.cell(row=9, column=args.count + 1).value = 0
    # np.random.seed(args.count+1)
    # labels = np.random.randint(low=0, high=5610,size = 1000)
    # X_test2 = x_test1[labels]
    # Y_test2 = y_test1[labels]

    # X_test2 = np.vstack((x_test1, x_test2))
    # Y_test2 = np.vstack((y_test1, y_test2))

    print(X_test2.shape)
    print(Y_test2.shape)

    get_score(X_test2,Y_test2,model,args.count+1)
    # get_score_driving(X_test2,Y_test2,model,args.count+1)

    target_lsa = fetch_lsa(model=model, x_train=x_train, x_target=X_test2, target_name=args.target, layer_names=layer_names, args=args)
    target_cov = get_sc(lower=args.lsa_lower, upper=args.lsa_upper, k=args.n_bucket, sa=target_lsa)

    target_lsa = [x for x in target_lsa if str(x) != 'nan']
    # sheet.cell(row=2, column=args.count+1).value = np.max(target_lsa)
    # sheet.cell(row=3, column=args.count + 1).value = np.min(target_lsa)
    sheet.cell(row=4, column=args.count + 1).value = target_cov
    print(len(target_lsa))
    print("max", np.max(target_lsa))
    print("min", np.min(target_lsa))
    print(infog("{} coverage: ".format(args.target) + str(target_cov)))


    target_dsa = fetch_dsa(model, x_train, X_test2, args.target, layer_names, args)
    target_cov = get_sc(lower=args.dsa_lower, upper=args.dsa_upper, k=args.n_bucket, sa=target_dsa)

    # sheet.cell(row=5, column=args.count+1).value = np.max(target_dsa)
    # sheet.cell(row=6, column=args.count + 1).value = np.min(target_dsa)
    sheet.cell(row=7, column=args.count + 1).value = target_cov
    print(infog("{} coverage: ".format(args.target) + str(target_cov)))
    print(len(target_dsa))
    print("max", np.max(target_dsa))
    print("min", np.min(target_dsa))

    workbook.save(os.path.join(basedir, "result","{}-{}samples-500groups-sc-{}.xlsx".format(args.d, args.size, args.target)))
    workbook.close()


    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)