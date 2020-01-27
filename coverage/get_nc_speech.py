from tools.utils4 import NeuronCoverage,K_SectionNeuronCoverage
from keras.datasets import cifar10, cifar100, fashion_mnist
import keras
import openpyxl
import os
import argparse
import numpy as np

from keras.datasets import mnist
from keras.models import load_model, Model

import datetime
from keras.applications import vgg19,resnet50

import deepspeech_utils as ds_utils
from deepspeech_utils import DSDataUtils
from deepspeech_utils import dataGen_mfcc_ctc
from deepspeech_text import Alphabet
import keras.backend as K

workbook = openpyxl.Workbook()
sheet = workbook.active

CLIP_MIN = -0.5
CLIP_MAX = 0.5

def get_score(x_test, y_test, model, column, row):
    total = 0
    diff = 0
    correct = 0
    alphabet = Alphabet(os.path.join(basedir,'data/alphabet.txt'))

    y_true = [[alphabet._label_to_str[y] for y in x] for x in y_test]
    y_true = [''.join(x).strip() for x in y_true]

    y_pred = model.predict(x_test, verbose=2)
    input_len = np.ones(y_pred.shape[0]) * y_pred.shape[1]

    y_pred = K.get_value(K.ctc_decode(y_pred, input_length=input_len)[0][0])
    y_pred = [[alphabet._label_to_str[y] for y in x if y >= 0] for x in y_pred]
    y_pred = [''.join(x).strip() for x in y_pred]
    y_pred = [DSDataUtils.get_bestmatch_keywords_using_wer(x) for x in y_pred]

    for a, b in list(zip(y_true, y_pred)):
        total += 1
        if a == b:
            correct += 1

    # print(f'accurary {round(correct / total, 3)} {correct} correct predictions in all {total} inputs')
    print(correct / total)
    print(correct)
    print(total)

    sheet.cell(row=row, column=column).value = (1 - correct / total)
    return correct / total

if __name__=="__main__":
    start = datetime.datetime.now()

    basedir = os.path.abspath(os.path.dirname(__file__))

    parser = argparse.ArgumentParser()
    parser.add_argument("--size", "-size", help="the number of each group", type=int, default=1000)
    parser.add_argument("--d", "-d", help="Dataset", type=str, default="speech")

    # parser.add_argument(
    #     "--num_classes",
    #     "-num_classes",
    #     help="The number of classes",
    #     type=int,
    #     default=10,
    # )
    parser.add_argument(
        "--target",
        "-target",
        help="Target input set (test or adversarial set)",
        type=str,
        default="cw",
    )

    args = parser.parse_args()
    print(args)

    batch_size = 6471
    testGen = dataGen_mfcc_ctc(ds_utils.testfile, batchSize=batch_size)
    [x_test1, y_test1, _, _], _ = next(DSDataUtils.gen_ctc_byclass(testGen))

    advfile = os.path.join(basedir,'data/speechcmd_adv.csv')
    batch_size = 6824
    advGen = dataGen_mfcc_ctc(advfile, batchSize=batch_size)
    [x_test2, y_test2, _, _], _ = next(DSDataUtils.gen_ctc_byclass(advGen))

    model = load_model("./model/deepspeech-speech-command_origin.h5")
    model.summary()

    import pickle
    with open('labels-500groups-1000samples.pk', 'rb') as fp:
        labels = pickle.load(fp)

    with open("boundry_dict_{}".format(args.d), "rb")as file:
        boundry = pickle.load(file)

    for count in range(500):
        p = labels['p'][count]
        p_100 = int(100*p)
        p_500 = int(500*p)

        start2 = datetime.datetime.now()

        X_test_100 = np.vstack((x_test1[labels['normal'][count][:100-p_100]], x_test2[labels['adv'][count][:p_100]]))
        Y_test_100 = np.vstack((y_test1[labels['normal'][count][:100-p_100]], y_test2[labels['adv'][count][:p_100]]))

        X_test_400 = np.vstack((x_test1[labels['normal'][count][100-p_100 : 500-p_500]], x_test2[labels['adv'][count][p_100 : p_500]]))
        Y_test_400 = np.vstack((y_test1[labels['normal'][count][100-p_100 : 500-p_500]], y_test2[labels['adv'][count][p_100 : p_500]]))

        X_test_500 = np.vstack((x_test1[labels['normal'][count][500-p_500:]],
                                x_test2[labels['adv'][count][p_500:]]))
        Y_test_500 = np.vstack((y_test1[labels['normal'][count][500-p_500:]],
                                y_test2[labels['adv'][count][p_500:]]))

        X_test2 = np.vstack((X_test_100, X_test_400))
        Y_test2 = np.vstack((Y_test_100, Y_test_400))

        s = get_score(X_test_100, Y_test_100, model, count + 1, 22)
        s = get_score(X_test2, Y_test2, model, count + 1, 23)

        X_test2 = np.vstack((X_test2, X_test_500))
        Y_test2 = np.vstack((Y_test2, Y_test_500))

        # X_test2 = np.vstack((x_test1[labels['normal'][count]], x_test2[labels['adv'][count]]))
        # Y_test2 = np.vstack((y_test1[labels['normal'][count]], y_test2[labels['adv'][count]]))

        print(count)

        sheet.cell(row=1, column=count+1).value = count
        sheet.cell(row=20, column=count+1).value = labels['p'][count]

        # 单纯的adv或者normal，不混合
        # sheet.cell(row=20, column=count + 1).value = 0
        # import random
        # labels = random.sample(range(0,9999),5000)
        # labels = np.random.choice(a=6471, size=1000, replace=False, p=None)
        # labels = np.random.randint(low=0, high=6471,size = 1000)
        # X_test2 = x_test1[labels]
        # Y_test2 = y_test1[labels]

        # X_test2 = np.vstack((X_test2, x_test2[labels]))
        # Y_test2 = np.vstack((Y_test2, y_test2[labels]))

        # X_test2 = x_test1
        # Y_test2 = y_test1

        # s = get_score(X_test2[:100], Y_test2[:100], model, count + 1, 22)
        # s = get_score(X_test2[:500], Y_test2[:500], model, count + 1, 23)

        print(X_test2.shape)
        print(Y_test2.shape)
        s = get_score(X_test2,Y_test2,model,count+1, 24)

        coverage1 = NeuronCoverage(model=model)
        final_coverage_rate = coverage1.final_coverage(inputs=X_test2, threshold=0.5, K=3, b=boundry)

        sheet.cell(row=3, column=count+1).value = final_coverage_rate[0]
        sheet.cell(row=4, column=count+1).value = final_coverage_rate[1]
        sheet.cell(row=5, column=count+1).value = final_coverage_rate[2]
        sheet.cell(row=6, column=count+1).value = final_coverage_rate[3]

        sheet.cell(row=7, column=count + 1).value = final_coverage_rate[4]
        sheet.cell(row=8, column=count + 1).value = final_coverage_rate[5]
        sheet.cell(row=9, column=count + 1).value = final_coverage_rate[6]
        sheet.cell(row=10, column=count + 1).value = final_coverage_rate[7]

        sheet.cell(row=11, column=count + 1).value = final_coverage_rate[8]
        sheet.cell(row=12, column=count + 1).value = final_coverage_rate[9]
        sheet.cell(row=13, column=count + 1).value = final_coverage_rate[10]
        sheet.cell(row=14, column=count + 1).value = final_coverage_rate[11]
        print("------------")

        start3 = datetime.datetime.now()
        coverage2 = K_SectionNeuronCoverage(model=model)
        kmnc = coverage2.final_coverage(inputs=X_test2, k_section=1000, b=boundry)
        elapsed3 = (datetime.datetime.now() - start3)
        print("Time used: ", elapsed3)
        print("------------")
        sheet.cell(row=16, column=count + 1).value = kmnc[0]
        sheet.cell(row=17, column=count + 1).value = kmnc[1]
        sheet.cell(row=18, column=count + 1).value = kmnc[2]


        workbook.save(os.path.join(basedir, "result","{}-{}samples-500groups-kmnc-{}.xlsx".format(args.d, args.size, args.target)))
        workbook.close()

        elapsed2 = (datetime.datetime.now() - start2)
        print("Time used: ", elapsed2)


    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)