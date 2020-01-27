from tools.utils4 import NeuronCoverage,K_SectionNeuronCoverage
from keras.datasets import cifar10, cifar100, fashion_mnist
import keras
import openpyxl
import os
import argparse
import numpy as np

from keras.datasets import mnist
from keras.models import load_model, Model

import datetime
from keras.applications import vgg19,resnet50
# import SVNH_DatasetUtil

workbook = openpyxl.Workbook()
sheet = workbook.active

CLIP_MIN = -0.5
CLIP_MAX = 0.5

def get_score(x_test, y_test, model, column, row):
    Y_test = keras.utils.to_categorical(y_test, args.num_classes)

    score = model.evaluate(x_test, Y_test, verbose=0)

    print('Test accuracy:', score[1])
    print('预测错的数目：', len(x_test)*(1-score[1]))

    sheet.cell(row=row, column=column).value = (1 - score[1])
    return score

def get_adv_imagenet(**kwargs):
    for i in range(0, 1000, 100):
        image_path = os.path.join(basedir, 'data', 'imagenet_adv/{}_imagenet_image_{}_resnet50.npy'.format(args.target, i))
        label_path = os.path.join(basedir, 'data', 'imagenet_adv/{}_imagenet_label_{}_resnet50.npy'.format(args.target, i))
        x_test = np.load(image_path)
        y_test = np.load(label_path)

        y = []
        for temp in y_test:
            temp_list = []
            temp_list.append(int(temp))
            y.append(temp_list)
        y = np.array(y)

        if i == 0:
            X_test = x_test
            Y_test = y
        else:
            X_test = np.vstack((X_test, x_test))
            Y_test = np.vstack((Y_test, y))

    return X_test,Y_test

def get_adv_mnist(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_mnist_image_lenet5.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_mnist_label_lenet5.npy'.format(args.target))
    x_test = np.load(image_path)
    x_test = x_test.astype("float32")

    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test,y

def get_adv_cifar10(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_cifar10_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_cifar10_label.npy'.format(args.target))
    x_test = np.load(image_path)

    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_svhn(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_svhn_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_svhn_label.npy'.format(args.target))
    x_test = np.load(image_path)

    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_cifar100(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_cifar100_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_cifar100_label.npy'.format(args.target))
    x_test = np.load(image_path)

    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

def get_adv_fashion(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/{}_fashion_image.npy'.format(args.target))
    label_path = os.path.join(basedir,'data','adv_image/{}_fashion_label.npy'.format(args.target))
    x_test = np.load(image_path)

    y_test = np.load(label_path)

    y = []
    for temp in y_test:
        temp_list = []
        temp_list.append(int(temp))
        y.append(temp_list)
    y = np.array(y)

    return x_test, y

if __name__=="__main__":
    start = datetime.datetime.now()

    basedir = os.path.abspath(os.path.dirname(__file__))

    parser = argparse.ArgumentParser()
    parser.add_argument("--size", "-size", help="the number of each group", type=int, default=1000)
    parser.add_argument("--d", "-d", help="Dataset", type=str, default="imagenet")

    parser.add_argument(
        "--num_classes",
        "-num_classes",
        help="The number of classes",
        type=int,
        default=1000,
    )
    parser.add_argument(
        "--target",
        "-target",
        help="Target input set (test or adversarial set)",
        type=str,
        default="cw",
    )

    args = parser.parse_args()
    print(args)

    if args.d == "mnist":
        (x_train, y_train), (x_test1, y_test) = mnist.load_data()

        (x_test2, y_test2) = get_adv_mnist()
        x_train = x_train.reshape(-1, 28, 28, 1)
        x_test1 = x_test1.reshape(-1, 28, 28, 1)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

        y_test1 = np.expand_dims(y_test,axis=1)

        # Load pre-trained model.
        model = load_model("./model/lenet-5.h5")
        # model = load_model("./model/model_mnist.hdf5")
        # model.summary()

    elif args.d == "cifar":
        (x_train, y_train), (x_test1, y_test1) = cifar10.load_data()

        (x_test2, y_test2) = get_adv_cifar10()

        model = load_model("./model/cifar10-vgg16_model_alllayers.h5")
        # model = load_model("./model/model_cifar10.h5")
        model.summary()

        x_train = x_train.reshape(x_train.shape[0], 32, 32, 3)
        x_test1 = x_test1.reshape(x_test1.shape[0], 32, 32, 3)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

    elif args.d == "svhn":
        (x_train, y_train), (x_test1, y_test1) = SVNH_DatasetUtil.load_data()
        (x_test2, y_test2) = get_adv_svhn()

        model = load_model("./model/model_svhn.hdf5")
        model.summary()
        y_test1 = np.argmax(y_test1, axis=1)
        y_test1 = y_test1.reshape(-1, 1)

        x_train = x_train.reshape(x_train.shape[0], 32, 32, 3)
        x_test1 = x_test1.reshape(x_test1.shape[0], 32, 32, 3)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

    elif args.d == "fashion":
        (x_train, y_train), (x_test1, y_test1) = fashion_mnist.load_data()

        (x_test2, y_test2) = get_adv_fashion()
        x_train = x_train.reshape(-1, 28, 28, 1)
        x_test1 = x_test1.reshape(-1, 28, 28, 1)
        y_train = y_train.reshape(-1, 1)
        y_test1 = y_test1.reshape(-1, 1)

        model = load_model("./model/model_fashion.hdf5")
        model.summary()

    elif args.d == "cifar100":
        (x_train, y_train), (x_test1, y_test1) = cifar100.load_data()

        (x_test2, y_test2) = get_adv_cifar100()

        model = load_model("./model/model_cifar100.h5")
        model.summary()

        x_train = x_train.reshape(x_train.shape[0], 32, 32, 3)
        x_test1 = x_test1.reshape(x_test1.shape[0], 32, 32, 3)

        x_train = x_train.astype('float32')
        x_test1 = x_test1.astype('float32')

        x_train /= 255
        x_test1 /= 255

    elif args.d == "imagenet":
        data_path = os.path.join(basedir, 'data', "imagenet-val-5000.npz")
        # data_path = os.path.join(basedir, 'data', "imagenet.npz")
        data = np.load(data_path)
        x_test1, y_test1 = data['x_test'], data['y_test']

        # x_test1 = vgg19.preprocess_input(x_test1)
        x_test1 = resnet50.preprocess_input(x_test1)

        (x_test2, y_test2) = get_adv_imagenet()
        # x_test2 = vgg19.preprocess_input(x_test2)
        x_test2 = resnet50.preprocess_input(x_test2)
        # model = vgg19.VGG19(weights='imagenet')
        model = resnet50.ResNet50(weights='imagenet')
        model.summary()

        model.compile(loss="categorical_crossentropy", optimizer="adadelta", metrics=["accuracy"])
        y_test1 = y_test1.reshape(y_test1.shape[0],1)


    import pickle
    with open('labels-500groups-1000samples.pk', 'rb') as fp:
        labels = pickle.load(fp)

    with open("boundry_dict_{}_resnet50".format(args.d), "rb")as file:
        boundry = pickle.load(file)

    # with open("boundry_dict_cifar10_vgg16", "rb")as file:
    #     boundry = pickle.load(file)

    for count in range(500):
        p = labels['p'][count]
        p_100 = int(100*p)
        p_500 = int(500*p)

        start2 = datetime.datetime.now()

        X_test_100 = np.vstack((x_test1[labels['normal'][count][:100-p_100]], x_test2[labels['adv'][count][:p_100]]))
        Y_test_100 = np.vstack((y_test1[labels['normal'][count][:100-p_100]], y_test2[labels['adv'][count][:p_100]]))

        X_test_400 = np.vstack((x_test1[labels['normal'][count][100-p_100 : 500-p_500]], x_test2[labels['adv'][count][p_100 : p_500]]))
        Y_test_400 = np.vstack((y_test1[labels['normal'][count][100-p_100 : 500-p_500]], y_test2[labels['adv'][count][p_100 : p_500]]))

        X_test_500 = np.vstack((x_test1[labels['normal'][count][500-p_500:]],
                                x_test2[labels['adv'][count][p_500:]]))
        Y_test_500 = np.vstack((y_test1[labels['normal'][count][500-p_500:]],
                                y_test2[labels['adv'][count][p_500:]]))

        X_test2 = np.vstack((X_test_100, X_test_400))
        Y_test2 = np.vstack((Y_test_100, Y_test_400))

        s = get_score(X_test_100, Y_test_100, model, count + 1, 22)
        s = get_score(X_test2, Y_test2, model, count + 1, 23)

        X_test2 = np.vstack((X_test2, X_test_500))
        Y_test2 = np.vstack((Y_test2, Y_test_500))

        # X_test2 = np.vstack((x_test1[labels['normal'][count]], x_test2[labels['adv'][count]]))
        # Y_test2 = np.vstack((y_test1[labels['normal'][count]], y_test2[labels['adv'][count]]))

        print(count)
        # print(X_test2.shape)
        # print(Y_test2.shape)

        sheet.cell(row=1, column=count+1).value = count
        sheet.cell(row=20, column=count+1).value = labels['p'][count]

        # 单纯的adv或者normal，不混合
        # sheet.cell(row=20, column=count + 1).value = 0
        # labels = np.random.randint(low=0, high=50000,size = 1000)
        # X_test2 = x_test1[labels]
        # Y_test2 = y_test1[labels]
        #
        # X_test2 = np.vstack((X_test2, x_test2[labels]))
        # Y_test2 = np.vstack((Y_test2, y_test2[labels]))

        # X_test2 = x_test1
        # Y_test2 = y_test1

        # s = get_score(X_test2[:100], Y_test2[:100], model, count + 1, 22)
        # s = get_score(X_test2[:500], Y_test2[:500], model, count + 1, 23)
        s = get_score(X_test2,Y_test2,model,count+1, 24)


        coverage1 = NeuronCoverage(model=model)
        final_coverage_rate = coverage1.final_coverage(inputs=X_test2, threshold=0.5, K=3, b=boundry)

        sheet.cell(row=3, column=count+1).value = final_coverage_rate[0]
        sheet.cell(row=4, column=count+1).value = final_coverage_rate[1]
        sheet.cell(row=5, column=count+1).value = final_coverage_rate[2]
        sheet.cell(row=6, column=count+1).value = final_coverage_rate[3]

        sheet.cell(row=7, column=count + 1).value = final_coverage_rate[4]
        sheet.cell(row=8, column=count + 1).value = final_coverage_rate[5]
        sheet.cell(row=9, column=count + 1).value = final_coverage_rate[6]
        sheet.cell(row=10, column=count + 1).value = final_coverage_rate[7]

        sheet.cell(row=11, column=count + 1).value = final_coverage_rate[8]
        sheet.cell(row=12, column=count + 1).value = final_coverage_rate[9]
        sheet.cell(row=13, column=count + 1).value = final_coverage_rate[10]
        sheet.cell(row=14, column=count + 1).value = final_coverage_rate[11]
        print("------------")

        start3 = datetime.datetime.now()
        coverage2 = K_SectionNeuronCoverage(model=model)
        kmnc = coverage2.final_coverage(inputs=X_test2, k_section=1000, b=boundry)
        elapsed3 = (datetime.datetime.now() - start3)
        print("Time used: ", elapsed3)
        print("------------")
        sheet.cell(row=16, column=count + 1).value = kmnc[0]
        sheet.cell(row=17, column=count + 1).value = kmnc[1]
        sheet.cell(row=18, column=count + 1).value = kmnc[2]


        workbook.save(os.path.join(basedir, "result","{}-{}samples-500groups-kmnc-{}-resnet50.xlsx".format(args.d, args.size, args.target)))
        workbook.close()

        elapsed2 = (datetime.datetime.now() - start2)
        print("Time used: ", elapsed2)


    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)